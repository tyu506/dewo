# tools_hf.py
# -*- coding: utf-8 -*-
"""
smolagents 基线/通用编排所用的 Hugging Face 工具层（Tools）

设计目标（尽量“不过度设计”，但能稳定跑实验）：
1) 对外只暴露少量工具（用于 baseline 的可比性与稳定性）：
   - search_models / get_model_card / get_model_info / inspect_task / infer / get_file_info
2) 内部复用 hf_client.py 的 ModelInferenceClient（底层 Hub + Inference Providers 封装）
3) 工具输出“可 JSON 序列化”
   - 文本/数字/list/dict：直接返回
   - 二进制（image/audio/video bytes 或 PIL 图像）：自动落盘，返回 {"type": "...", "path": "..."} 或 path
4) 记录最小的工具统计（调用次数/耗时/错误数），便于后续算 API Call Count、latency 分解

注意：
- 为了公平对比，建议 baseline agent 初始化时 add_base_tools=False，不引入 websearch 等额外工具能力。
- token 默认从环境变量 HF_TOKEN 读取，也可在调用 infer 时显式传入 hf_token。
"""

from __future__ import annotations

import base64
import csv
import concurrent.futures
import json
import mimetypes
import os
import time
import traceback
import uuid
import threading
from datetime import datetime, timezone, date
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from smolagents import tool

from app.utils.output_viz import prepare_infer_result_for_ui

from .hf_clients import ModelInferenceClient, TASK_TO_PIPELINE, TASK_CONFIGS

# 说明：为了兼容部分 provider（如 DeepSeek）的 function calling JSON schema 校验，
# infer 工具的 parameters 只允许 JSON 标量类型，避免出现 "type": "any" 的 schema。
JSONScalar = Union[str, int, float, bool]


def _env_int(name: str, default: Optional[int] = None) -> Optional[int]:
    v = os.environ.get(name)
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    try:
        return int(float(s))
    except Exception:
        return default


def _env_float(name: str, default: Optional[float] = None) -> Optional[float]:
    v = os.environ.get(name)
    if v is None:
        return default
    s = str(v).strip()
    if not s:
        return default
    try:
        return float(s)
    except Exception:
        return default


def _clamp_int(x: Any, lo: int, hi: int, default: int) -> int:
    try:
        v = int(float(x))
    except Exception:
        v = int(default)
    return max(int(lo), min(int(v), int(hi)))


def _hard_limit_search_models_limit_max() -> Optional[int]:
    """
    工具级硬上限：search_models(limit) 最大允许值。
    runner 应在启动时从 config.yaml/baseline_budget.K 注入：
      - SEARCH_MODELS_LIMIT_MAX
    """
    v = _env_int("SEARCH_MODELS_LIMIT_MAX", default=None)
    if v is None:
        return None
    return max(1, int(v))


def _hard_limit_infer_timeout_max_s() -> Optional[float]:
    """
    工具级硬上限：infer(timeout_s) 最大允许值（秒）。
    runner 应在启动时从 config.yaml/baseline_budget.infer_timeout_s 注入：
      - INFER_TIMEOUT_MAX_S
    """
    v = _env_float("INFER_TIMEOUT_MAX_S", default=None)
    if v is None:
        return None
    return max(1.0, float(v))


def _hard_limit_model_card_max_chars_max() -> Optional[int]:
    """
    工具级硬上限：get_model_card(max_chars) 最大允许值。
    可选注入：
      - MODEL_CARD_MAX_CHARS_MAX
    """
    v = _env_int("MODEL_CARD_MAX_CHARS_MAX", default=None)
    if v is None:
        return None
    return max(100, int(v))


def _looks_like_timeout_error(exc: BaseException) -> bool:
    """最佳努力判断异常是否属于超时类错误。"""
    name = type(exc).__name__.lower()
    msg = str(exc).lower()
    timeout_tokens = (
        "timeout",
        "timed out",
        "read timeout",
        "connect timeout",
        "deadline exceeded",
    )
    if "timeout" in name:
        return True
    return any(tok in msg for tok in timeout_tokens)


# =========================
# 全局：工具统计（线程安全）
# =========================
# 作用：记录各个工具（search_models / get_model_card / get_model_info / inspect_task /
#      infer / get_file_info）的调用次数、总耗时、错误次数和最后一次错误信息，
# 供 runner 写 JSONL 时写入 RunRecord.tool_stats，用于算 API Call Count、latency 分解等评测指标。
# 多线程/多 agent 可能同时写统计，所以用锁保证数据一致。

_STATS_LOCK = threading.Lock()  # 写 _TOOL_STATS 时加锁，避免竞态
_TOOL_STATS: Dict[str, Dict[str, Any]] = {
    "counts": {
        "search_models": 0,
        "get_model_card": 0,
        "inspect_task": 0,
        "infer": 0,
        "get_file_info": 0,
        "get_model_info": 0,
    },  # 各工具调用次数
    "durations_ms": {
        "search_models": 0.0,
        "get_model_card": 0.0,
        "inspect_task": 0.0,
        "infer": 0.0,
        "get_file_info": 0.0,
        "get_model_info": 0.0,
    },  # 各工具累计耗时（毫秒）
    "errors": {
        "search_models": 0,
        "get_model_card": 0,
        "inspect_task": 0,
        "infer": 0,
        "get_file_info": 0,
        "get_model_info": 0,
    },  # 各工具失败次数
    "error_list": {
        "search_models": [],
        "get_model_card": [],
        "inspect_task": [],
        "infer": [],
        "get_file_info": [],
        "get_model_info": [],
    },  # 各工具的错误信息列表（按发生顺序记录完整错误）
    # 记录每个工具“最后一次调用”是否成功，用于更严格的成功判定（如 MC1）
    "last_ok": {
        "search_models": None,
        "get_model_card": None,
        "inspect_task": None,
        "infer": None,
        "get_file_info": None,
        "get_model_info": None,
    },  # Optional[bool]: last call status
}

# =========================
# 媒体落盘记录（用于 runner 回填 output_file）
# =========================
_MEDIA_SAVED_LOCK = threading.Lock()
_MEDIA_SAVED: List[Dict[str, Any]] = []


def reset_saved_media() -> None:
    """清空本次样本的媒体落盘记录（仅用于 runner/评测，不影响 tool stats）。"""
    global _MEDIA_SAVED
    with _MEDIA_SAVED_LOCK:
        _MEDIA_SAVED = []


def get_saved_media() -> List[Dict[str, Any]]:
    """获取本次样本已落盘媒体记录副本（每项至少包含 type/path）。"""
    with _MEDIA_SAVED_LOCK:
        return list(_MEDIA_SAVED)


def _record_saved_media(*, output_type: str, path: str, method: str) -> None:
    """记录一次媒体落盘，用于 runner 回填 output_file。"""
    try:
        with _MEDIA_SAVED_LOCK:
            _MEDIA_SAVED.append(
                {
                    "type": str(output_type),
                    "path": str(path),
                    "method": str(method),
                    "ts": time.time(),
                }
            )
    except Exception:
        # 记录失败不应影响主流程
        return


# =========================
# 全局：工具调用 trace（中断也能留痕）
# =========================
# 作用：将每次工具调用的参数/结果/错误以 JSONL 追加写入文件。
# - runner 可通过环境变量 TOOL_TRACE_PATH 为“每条样本”指定一个 trace 文件路径
# - 发生 KeyboardInterrupt 或进程中断时，已写入的 trace 仍可用于复盘
_TRACE_LOCK = threading.Lock()


def _truncate_for_trace(obj: Any, n: int = 2000) -> Any:
    if isinstance(obj, str):
        return obj if len(obj) <= n else obj[:n] + "...(truncated)"
    if isinstance(obj, dict):
        return {k: _truncate_for_trace(v, n) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_truncate_for_trace(v, n) for v in obj]
    return str(obj)


def _trace_path() -> Optional[str]:
    p = os.environ.get("TOOL_TRACE_PATH")
    if not p or not str(p).strip():
        return None
    return str(p)


def _append_trace(event: Dict[str, Any]) -> None:
    p = _trace_path()
    if p is None:
        return
    try:
        os.makedirs(os.path.dirname(p) or ".", exist_ok=True)
        line = json.dumps(event, ensure_ascii=False)
        with _TRACE_LOCK:
            with open(p, "a", encoding="utf-8") as f:
                f.write(line + "\n")
                f.flush()
    except Exception:
        # trace 不能影响主流程
        return


def _redact_args(d: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(d)
    for k in ("hf_token", "token", "api_key"):
        if k in out:
            out[k] = "***redacted***"
    return out


def reset_tool_stats() -> None:
    """重置工具统计（非 tool，仅供 runner/调试使用）。例如每条样本跑完后清零，避免和上一条样本的统计混在一起。"""
    with _STATS_LOCK:
        for k in _TOOL_STATS["counts"]:
            _TOOL_STATS["counts"][k] = 0
        for k in _TOOL_STATS["durations_ms"]:
            _TOOL_STATS["durations_ms"][k] = 0.0
        for k in _TOOL_STATS["errors"]:
            _TOOL_STATS["errors"][k] = 0
        for k in _TOOL_STATS["error_list"]:
            _TOOL_STATS["error_list"][k] = []
        for k in _TOOL_STATS["last_ok"]:
            _TOOL_STATS["last_ok"][k] = None


def get_tool_stats() -> Dict[str, Any]:
    """获取工具统计快照（非 tool，仅供 runner/调试使用）。返回拷贝，防止调用方误改影响内部状态。"""
    with _STATS_LOCK:
        # 通过 json 序列化再反序列化得到深拷贝，避免外部误改
        return json.loads(json.dumps(_TOOL_STATS, ensure_ascii=False))


def _stat_ok(name: str, dt_ms: float) -> None:
    """记录一次工具调用成功：次数+1，累计耗时累加 dt_ms。供各 @tool 在 return 前调用。"""
    with _STATS_LOCK:
        _TOOL_STATS["counts"][name] += 1
        _TOOL_STATS["durations_ms"][name] += float(dt_ms)
        try:
            _TOOL_STATS["last_ok"][name] = True
        except Exception:
            pass


def _stat_err(name: str, dt_ms: float, err: Exception) -> None:
    """记录一次工具调用失败：次数+1、耗时累加、错误数+1，并将完整错误追加到 error_list。"""
    with _STATS_LOCK:
        _TOOL_STATS["counts"][name] += 1
        _TOOL_STATS["durations_ms"][name] += float(dt_ms)
        _TOOL_STATS["errors"][name] += 1
        msg = f"{type(err).__name__}: {str(err)}"
        try:
            _TOOL_STATS["error_list"][name].append(msg)
        except Exception:
            # 统计失败不应影响主流程
            pass
        try:
            _TOOL_STATS["last_ok"][name] = False
        except Exception:
            pass


# =========================
# 全局：二进制输出落盘
# =========================
# 作用：推理结果可能是 image/audio/video 的 bytes 或 PIL 图像，不能直接写进 JSON/JSONL。
# 这里把这类结果保存到本地文件，返回 {"type": "image"|"audio"|"video", "path": "..."}，保证工具返回值始终可 JSON 序列化。

def _assets_dir() -> str:
    """
    返回媒体文件保存的目录路径。若目录不存在会自动创建。
    - 默认：outputs/assets
    - 可用环境变量 TOOL_ASSETS_DIR 覆盖（便于部署时指定磁盘或挂载路径）
    """
    p = os.getenv("TOOL_ASSETS_DIR", os.path.join("outputs", "assets"))
    os.makedirs(p, exist_ok=True)
    return p


def _guess_ext(output_type: str) -> str:
    """根据输出类型（来自 TASK_CONFIGS.output_type）猜测文件扩展名，用于落盘时的文件名。"""
    output_type = (output_type or "").lower()
    if output_type == "image":
        return ".png"
    if output_type == "audio":
        return ".wav"
    if output_type == "video":
        return ".mp4"
    return ".bin"


def _save_bytes(data: bytes, output_type: str) -> str:
    """把二进制数据写入 _assets_dir() 下的文件，文件名带时间戳和 uuid 避免冲突，返回完整路径。"""
    filename = f"{int(time.time())}_{uuid.uuid4().hex}{_guess_ext(output_type)}"
    path = os.path.join(_assets_dir(), filename)
    with open(path, "wb") as f:
        f.write(data)
    _record_saved_media(output_type=output_type, path=path, method="bytes")
    return path


def _maybe_save_media(result: Any, output_type_hint: str) -> Any:
    """
    把推理结果统一变成“可 JSON 序列化”的返回值，避免写 JSONL 时因二进制或 PIL 对象报错。
    策略：
    - bytes/bytearray -> 落盘，返回 {"type": output_type_hint, "path": 路径}
    - 有 .save 方法的对象（如 PIL Image）-> 保存为 png，返回 {"type": "image", "path": 路径}
    - dict/list/str/int/float/bool/None -> 原样返回
    - 其它 -> str(result) 兜底
    """
    # 1) 原始字节：保存到文件，返回 type+path，方便评测/前端按路径加载
    if isinstance(result, (bytes, bytearray)):
        path = _save_bytes(bytes(result), output_type_hint)
        return {"type": output_type_hint, "path": path}

    # 2) PIL 或类似“可 save 的对象”：不直接 import PIL，用 duck-typing 检测 .save，保存为 png
    if hasattr(result, "save") and callable(getattr(result, "save")):
        filename = f"{int(time.time())}_{uuid.uuid4().hex}.png"
        path = os.path.join(_assets_dir(), filename)
        try:
            result.save(path)
            _record_saved_media(output_type="image", path=path, method="PIL.save")
            return {"type": "image", "path": path}
        except Exception:
            return str(result)

    # 3) 已是可序列化类型：直接返回，不落盘
    if isinstance(result, (dict, list, str, int, float, bool)) or result is None:
        return result

    # 4) 其它类型（如自定义对象）：转成字符串，避免 json.dumps 报错
    return str(result)


def _to_jsonable(x: Any) -> Any:
    """
    将 huggingface_hub / pydantic / dataclass 等对象尽量转换为可 JSON 序列化的结构。
    """
    if x is None or isinstance(x, (str, int, float, bool)):
        return x
    if isinstance(x, (datetime, date)):
        return x.isoformat()
    if isinstance(x, Path):
        return str(x)
    if isinstance(x, Enum):
        return x.value
    if isinstance(x, dict):
        return {str(k): _to_jsonable(v) for k, v in x.items()}
    if isinstance(x, (list, tuple, set)):
        return [_to_jsonable(v) for v in x]

    # 常见：pydantic v2 / v1
    if hasattr(x, "model_dump") and callable(getattr(x, "model_dump")):
        return _to_jsonable(x.model_dump())
    if hasattr(x, "dict") and callable(getattr(x, "dict")):
        return _to_jsonable(x.dict())

    # dataclass 或一般对象
    if hasattr(x, "__dict__"):
        return _to_jsonable(vars(x))
    return str(x)


# =========================
# 文件信息/多模态资产解析
# =========================

_IMAGE_EXTS = {
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".gif",
    ".bmp",
    ".tiff",
    ".tif",
    ".heic",
    ".heif",
}
_AUDIO_EXTS = {
    ".wav",
    ".mp3",
    ".flac",
    ".ogg",
    ".opus",
    ".m4a",
    ".aac",
    ".wma",
}
_TABLE_EXTS = {
    ".csv",
    ".tsv",
    ".xlsx",
    ".xls",
}


def _ts_iso_utc(ts: float) -> str:
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()


def _safe_int(x: Any) -> Optional[int]:
    try:
        return int(x)
    except Exception:
        return None


def _detect_file_kind_by_suffix(path: str) -> str:
    ext = os.path.splitext(path)[1].lower()
    if ext in _IMAGE_EXTS:
        return "image"
    if ext in _AUDIO_EXTS:
        return "audio"
    if ext in _TABLE_EXTS:
        return "table"
    # mimetype 兜底
    mime, _ = mimetypes.guess_type(path)
    if mime:
        if mime.startswith("image/"):
            return "image"
        if mime.startswith("audio/"):
            return "audio"
        if mime in ("text/csv", "application/vnd.ms-excel"):
            return "table"
    return "unknown"


def _image_meta_best_effort(path: str) -> Dict[str, Any]:
    """
    图片元信息：优先 Pillow；无 Pillow 时对 PNG/JPEG 做轻量解析拿到宽高，其它仅返回基础信息。
    返回值必须可 JSON 序列化。
    """
    meta: Dict[str, Any] = {"format": None, "width": None, "height": None, "mode": None}

    # 1) Pillow（可选）
    try:
        from PIL import Image  # type: ignore

        with Image.open(path) as img:
            meta.update(
                {
                    "format": getattr(img, "format", None),
                    "width": _safe_int(getattr(img, "width", None)),
                    "height": _safe_int(getattr(img, "height", None)),
                    "mode": getattr(img, "mode", None),
                }
            )
            return meta
    except Exception:
        pass

    # 2) 无依赖兜底：解析 PNG/JPEG header 拿宽高
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext == ".png":
            with open(path, "rb") as f:
                sig = f.read(8)
                if sig != b"\x89PNG\r\n\x1a\n":
                    return meta
                _len = f.read(4)
                ctype = f.read(4)
                if ctype != b"IHDR":
                    return meta
                ihdr = f.read(13)
                if len(ihdr) == 13:
                    w = int.from_bytes(ihdr[0:4], "big")
                    h = int.from_bytes(ihdr[4:8], "big")
                    meta.update({"format": "PNG", "width": w, "height": h})
                    return meta
        if ext in (".jpg", ".jpeg"):
            with open(path, "rb") as f:
                if f.read(2) != b"\xff\xd8":
                    return meta
                while True:
                    b = f.read(1)
                    if not b:
                        break
                    if b != b"\xff":
                        continue
                    marker = f.read(1)
                    if not marker:
                        break
                    # 跳过填充
                    while marker == b"\xff":
                        marker = f.read(1)
                        if not marker:
                            break
                    if not marker:
                        break
                    # SOF0/1/2/3/5/6/7/9/A/B/D/E/F
                    if marker in (
                        b"\xc0",
                        b"\xc1",
                        b"\xc2",
                        b"\xc3",
                        b"\xc5",
                        b"\xc6",
                        b"\xc7",
                        b"\xc9",
                        b"\xca",
                        b"\xcb",
                        b"\xcd",
                        b"\xce",
                        b"\xcf",
                    ):
                        seg_len = int.from_bytes(f.read(2), "big")
                        data = f.read(seg_len - 2)
                        if len(data) >= 7:
                            h = int.from_bytes(data[1:3], "big")
                            w = int.from_bytes(data[3:5], "big")
                            meta.update({"format": "JPEG", "width": w, "height": h})
                        return meta
                    else:
                        seg_len_b = f.read(2)
                        if len(seg_len_b) != 2:
                            break
                        seg_len = int.from_bytes(seg_len_b, "big")
                        if seg_len < 2:
                            break
                        f.seek(seg_len - 2, 1)
    except Exception:
        return meta
    return meta


def _audio_meta_best_effort(path: str) -> Dict[str, Any]:
    """
    音频元信息：WAV 用标准库 wave 拿到时长/采样率等；其它格式返回基础信息并尽量给出可推断字段（可能为空）。
    """
    meta: Dict[str, Any] = {
        "format": None,
        "duration_s": None,
        "sample_rate_hz": None,
        "channels": None,
        "sample_width_bytes": None,
        "n_frames": None,
    }

    ext = os.path.splitext(path)[1].lower()
    if ext == ".wav":
        try:
            import wave

            with wave.open(path, "rb") as wf:
                fr = wf.getframerate()
                nf = wf.getnframes()
                ch = wf.getnchannels()
                sw = wf.getsampwidth()
                dur = (nf / float(fr)) if fr else None
                meta.update(
                    {
                        "format": "WAV",
                        "duration_s": float(dur) if dur is not None else None,
                        "sample_rate_hz": int(fr) if fr is not None else None,
                        "channels": int(ch) if ch is not None else None,
                        "sample_width_bytes": int(sw) if sw is not None else None,
                        "n_frames": int(nf) if nf is not None else None,
                    }
                )
                return meta
        except Exception:
            return meta

    # 其它音频格式：无额外依赖时难以可靠解析时长/码率，保持结构化字段但可为空
    meta["format"] = ext.lstrip(".").upper() if ext else None
    return meta


def _read_text_table_rows(path: str, return_table: bool = True) -> Dict[str, Any]:
    """
    读取 CSV/TSV 内容，返回：
    - meta: delimiter/encoding/rows/cols/columns
    - table: 完整二维数组（list[list[str]]），仅当 return_table=True 时返回（否则为 []）
    """
    ext = os.path.splitext(path)[1].lower()
    delimiter = "\t" if ext == ".tsv" else ","
    tried_encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"]
    last_err: Optional[str] = None
    rows: List[List[str]] = []
    used_encoding: Optional[str] = None

    for enc in tried_encodings:
        try:
            rows = []
            n_rows = 0
            n_cols_max = 0
            columns: List[str] = []
            with open(path, "r", encoding=enc, newline="") as f:
                reader = csv.reader(f, delimiter=delimiter)
                for r in reader:
                    rr = [str(x) for x in r]
                    if n_rows == 0:
                        columns = rr
                    n_rows += 1
                    n_cols_max = max(n_cols_max, len(rr))
                    if return_table:
                        rows.append(rr)
            used_encoding = enc
            last_err = None
            break
        except Exception as e:
            last_err = f"{type(e).__name__}: {str(e)}"
            continue

    if used_encoding is None:
        raise ValueError(f"无法读取表格文本文件，最后一次错误：{last_err}")

    return {
        "meta": {
            "format": "TSV" if ext == ".tsv" else "CSV",
            "encoding": used_encoding,
            "delimiter": delimiter,
            "n_rows": n_rows,
            "n_cols_max": n_cols_max,
            "columns": columns,
        },
        "table": rows if return_table else [],
    }


def _read_xlsx_table(
    path: str, sheet_name: Optional[str], return_table: bool = True
) -> Dict[str, Any]:
    """
    读取 XLSX 内容（默认第一个 sheet 或指定 sheet）。
    需要 openpyxl（若缺失则抛出可读错误信息）。
    """
    try:
        import openpyxl  # type: ignore
    except Exception as e:
        raise ImportError(
            "读取 .xlsx 需要依赖 openpyxl。请安装：pip install openpyxl"
        ) from e

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        sheets = list(wb.sheetnames)
        if not sheets:
            return {"meta": {"format": "XLSX", "sheets": []}, "table": []}
        use_sheet = sheet_name if (sheet_name in sheets) else sheets[0]
        ws = wb[use_sheet]
        rows: List[List[str]] = []
        n_rows = 0
        n_cols_max = 0
        columns: List[str] = []
        for row in ws.iter_rows(values_only=True):
            rr = ["" if v is None else str(v) for v in row]
            if n_rows == 0:
                columns = rr
            n_rows += 1
            n_cols_max = max(n_cols_max, len(rr))
            if return_table:
                rows.append(rr)
        return {
            "meta": {
                "format": "XLSX",
                "sheet_name": use_sheet,
                "sheets": sheets,
                "n_rows": n_rows,
                "n_cols_max": n_cols_max,
                "columns": columns,
            },
            "table": rows if return_table else [],
        }
    finally:
        try:
            wb.close()
        except Exception:
            pass


# =========================
# Tool 1: search_models
# =========================

@tool
def search_models(
    task_type: str,
    limit: int = 50,
    search: str = "",
    sort: str = "trending_score",
    direction: int = -1,
    warm_only: bool = True,
    hf_token: Optional[str] = None,
) -> Dict[str, Any]:
    """
    在 Hugging Face Hub 上按任务类型检索候选模型（用于在线模型发现）。

    Args:
        task_type: 你的 benchmark task_type（例如 "text_generation", "summarization", "translation", "object_detection" 等）。
        limit: 返回最多多少个候选模型。
        search: 可选的 free-text 搜索词（例如模型名/关键词）。
        sort: Hub 排序字段（常用 "trending_score" 或 "downloads" 或 "likes"等）。
        direction: 排序方向，-1 表示降序，1 表示升序。
        warm_only: 是否只检索 inference="warm" 的模型（更贴近“可执行性”评测）。
        hf_token: 可选，HF token；不传则从环境变量 HF_TOKEN 读取。

    Returns:
        一个 dict，包含候选模型的轻量信息：
        {
          "pipeline_tag": "...",
          "models": [
            {"model_id": "...", "pipeline_tag": "...", "tags": [...], "likes": 123, "downloads": 456},
            ...
          ]
        }
    """
    t0 = time.time()
    try:
        _append_trace(
            {
                "ts": time.time(),
                "tool": "search_models",
                "stage": "call",
                "args": _truncate_for_trace(_redact_args(locals())),
            }
        )

        # 工具级硬约束：search_models(limit) 上限（对 smolagents / ReAct 一视同仁）
        limit_max = _hard_limit_search_models_limit_max()
        if limit_max is not None:
            limit = _clamp_int(limit, lo=1, hi=limit_max, default=min(50, limit_max))

        client = ModelInferenceClient(token=hf_token, async_mode=False)
        models = client.list_models_by_task(
            task_type=task_type,
            limit=limit,
            sort=sort,
            search=search or None,
            direction=direction,
            warm_only=warm_only,
        )

        out: List[Dict[str, Any]] = []
        for m in models:
            # huggingface_hub 返回对象字段在不同版本略有差异，这里用 getattr 做鲁棒提取
            model_id = getattr(m, "modelId", None) or getattr(m, "id", None)
            out.append({
                "model_id": model_id,
                "pipeline_tag": getattr(m, "pipeline_tag", None),
                "tags": getattr(m, "tags", None),
                "likes": getattr(m, "likes", None),
                "downloads": getattr(m, "downloads", None),
            })

        dt_ms = (time.time() - t0) * 1000
        _stat_ok("search_models", dt_ms)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "search_models",
                "stage": "return",
                "ok": True,
                "dt_ms": dt_ms,
                "result": _truncate_for_trace(
                    {
                        "pipeline_tag": TASK_TO_PIPELINE.get(task_type),
                        "models_head": out[:5],
                        "models_count": len(out),
                    }
                ),
            }
        )

        return {
            "pipeline_tag": TASK_TO_PIPELINE.get(task_type),
            "models": out,
        }
    except Exception as e:
        dt_ms = (time.time() - t0) * 1000
        _stat_err("search_models", dt_ms, e)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "search_models",
                "stage": "error",
                "ok": False,
                "dt_ms": dt_ms,
                "error": f"{type(e).__name__}: {str(e)}",
            }
        )
        # tool 层直接抛错，让 agent 决定如何处理
        raise


# =========================
# Tool 2: get_model_card 
# =========================
@tool
def get_model_card(
    model_id: List[str],
    max_chars: int = 4000,
    hf_token: Optional[str] = None,
) -> Dict[str, Any]:
    """
    并行拉取多个模型的 model card 完整结构（JSON 化），仅对其中的 text 字段按 max_chars 做截断限制，最后整理并返回。

    Args:
        model_id: 模型 id 列表，例如 ["facebook/bart-large-cnn", "t5-small"]。
        max_chars: 每个 model card 的最大返回字符数（防止 prompt 膨胀）。
                   该参数设置了上限，若调用传入值大于上限，则自动截断为上限，否则使用调用传入值。
        hf_token: 可选，HF token；不传则从环境变量 HF_TOKEN 读取。

    Returns:
        dict，保证可 JSON 序列化，例如：
        {
          "n_total": 3,
          "n_ok": 2,
          "results": [
            {"model_id": "...", "ok": true, "model_card": "..."},
            {"model_id": "...", "ok": false, "error": "..."}
          ]
        }
    """
    t0 = time.time()
    try:
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_model_card",
                "stage": "call",
                "args": _truncate_for_trace(_redact_args(locals())),
            }
        )

        model_ids = [str(x) for x in (model_id or []) if str(x).strip()]
        if not model_ids:
            raise ValueError("model_id 不能为空；请传入至少 1 个模型 id。")

        # 工具级硬约束：get_model_card(max_chars) 上限（可选注入）
        max_chars_max = _hard_limit_model_card_max_chars_max()
        if max_chars_max is not None:
            # 仅在调用方 max_chars 超过环境上限时才生效（等同做“上限”）
            try:
                v = int(float(max_chars))
            except Exception:
                v = min(4000, max_chars_max)
            if v < 100:
                v = 100
            if v > max_chars_max:
                v = max_chars_max
            max_chars = v

        def _worker(mid: str) -> Dict[str, Any]:
            tt0 = time.time()
            try:
                client = ModelInferenceClient(token=hf_token, async_mode=False)
                card = client.get_model_card(model_id=mid)

                # 将完整结构转为 jsonable，并仅对顶层 text 字段应用 max_chars 截断
                card_json = _to_jsonable(card)
                if isinstance(card_json, dict):
                    text_val = card_json.get("text")
                    if isinstance(text_val, str) and max_chars is not None:
                        try:
                            mc = int(float(max_chars))
                        except Exception:
                            mc = 4000
                        if mc < 100:
                            mc = 100
                        if len(text_val) > mc:
                            card_json["text"] = text_val[:mc]

                dt_ms = (time.time() - tt0) * 1000
                return {
                    "model_id": mid,
                    "ok": True,
                    "model_card": card_json,
                    "dt_ms": float(dt_ms),
                }
            except Exception as e:
                dt_ms = (time.time() - tt0) * 1000
                return {
                    "model_id": mid,
                    "ok": False,
                    "error": f"{type(e).__name__}: {str(e)}",
                    "dt_ms": float(dt_ms),
                }

        max_workers = min(8, max(1, len(model_ids)))
        results: List[Dict[str, Any]] = []
        # 并行获取：每个 worker 自己创建 client，避免潜在线程安全问题
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(_worker, mid): mid for mid in model_ids}
            for fut in concurrent.futures.as_completed(futures):
                results.append(fut.result())

        # 按输入顺序整理返回（便于 agent/评测对齐）
        by_id = {r["model_id"]: r for r in results}
        ordered_results = [by_id[mid] for mid in model_ids if mid in by_id]

        n_ok = sum(1 for r in ordered_results if r.get("ok") is True)
        dt_ms = (time.time() - t0) * 1000

        _stat_ok("get_model_card", dt_ms)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_model_card",
                "stage": "return",
                "ok": True,
                "dt_ms": dt_ms,
                "result": _truncate_for_trace(
                    {
                        "n_total": len(model_ids),
                        "n_ok": n_ok,
                        "results_head": ordered_results[:2],
                    },
                    2000,
                ),
            }
        )

        return {
            "n_total": len(model_ids),
            "n_ok": int(n_ok),
            "results": ordered_results,
        }
    except Exception as e:
        dt_ms = (time.time() - t0) * 1000
        _stat_err("get_model_card", dt_ms, e)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_model_card",
                "stage": "error",
                "ok": False,
                "dt_ms": dt_ms,
                "error": f"{type(e).__name__}: {str(e)}",
            }
        )
        raise


# =========================
# Tool 3: get_model_info
# =========================

@tool
def get_model_info(
    model_id: List[str],
    expand: Optional[List[str]] = None,
    hf_token: Optional[str] = None,
) -> Dict[str, Any]:
    """
    并行拉取多个模型的 model info，返回完整 JSON 结构。

    Args:
        model_id: 模型 id 列表，例如 ["facebook/bart-large-cnn", "t5-small"]。
        expand: 传递给 huggingface_hub.HfApi.model_info 的 expand 参数（属性名列表），
                当为 None 时不使用 expand（即返回完整默认字段）；当为列表时，只返回该列表中的字段。
        hf_token: 可选，HF token；不传则从环境变量 HF_TOKEN 读取。

    Returns:
        dict，保证可 JSON 序列化，例如：
        {
          "n_total": 3,
          "n_ok": 2,
          "results": [
            {"model_id": "...", "ok": true, "info": {...}},
            {"model_id": "...", "ok": false, "error": "..."}
          ]
        }
    """
    t0 = time.time()
    try:
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_model_info",
                "stage": "call",
                "args": _truncate_for_trace(_redact_args(locals())),
            }
        )

        model_ids = [str(x) for x in (model_id or []) if str(x).strip()]
        if not model_ids:
            raise ValueError("model_id 不能为空；请传入至少 1 个模型 id。")

        def _worker(mid: str) -> Dict[str, Any]:
            tt0 = time.time()
            try:
                client = ModelInferenceClient(token=hf_token, async_mode=False)
                info = client.get_model_info(mid, expand=expand)
                info_json = _to_jsonable(info)
                dt_ms = (time.time() - tt0) * 1000
                return {
                    "model_id": mid,
                    "ok": True,
                    "info": info_json,
                    "dt_ms": float(dt_ms),
                }
            except Exception as e:
                dt_ms = (time.time() - tt0) * 1000
                return {
                    "model_id": mid,
                    "ok": False,
                    "error": f"{type(e).__name__}: {str(e)}",
                    "dt_ms": float(dt_ms),
                }

        max_workers = min(8, max(1, len(model_ids)))
        results: List[Dict[str, Any]] = []
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as ex:
            futures = {ex.submit(_worker, mid): mid for mid in model_ids}
            for fut in concurrent.futures.as_completed(futures):
                results.append(fut.result())

        by_id = {r["model_id"]: r for r in results}
        ordered_results = [by_id[mid] for mid in model_ids if mid in by_id]

        n_ok = sum(1 for r in ordered_results if r.get("ok") is True)
        dt_ms = (time.time() - t0) * 1000

        _stat_ok("get_model_info", dt_ms)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_model_info",
                "stage": "return",
                "ok": True,
                "dt_ms": dt_ms,
                "result": _truncate_for_trace(
                    {
                        "n_total": len(model_ids),
                        "n_ok": n_ok,
                        "results_head": ordered_results[:2],
                    },
                    2000,
                ),
            }
        )

        return {
            "n_total": len(model_ids),
            "n_ok": int(n_ok),
            "results": ordered_results,
        }
    except Exception as e:
        dt_ms = (time.time() - t0) * 1000
        _stat_err("get_model_info", dt_ms, e)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_model_info",
                "stage": "error",
                "ok": False,
                "dt_ms": dt_ms,
                "error": f"{type(e).__name__}: {str(e)}",
            }
        )
        raise


# =========================
# Tool 4: inspect_task
# =========================

@tool
def inspect_task(
    task_type: str,
    async_mode: bool = False,
    hf_token: Optional[str] = None,
) -> Dict[str, Any]:
    """
    查询指定 task_type 在当前 Hugging Face InferenceClient 上的调用结构信息。

    作用：
    - 核心作用：帮助 agent 在调用 infer 之前了解：infer工具底层实际映射到哪个方法（mapped_method）、需要哪些必填字段（required_args）、方法签名的参数列表及默认值等。
    - 实际推理时，请根据infer工具的parameters参数结构要求进行构造。
    - 这有助于减少因参数不匹配导致的推理失败，特别是不同任务之间的差异。

    Args:
        task_type: 任务类型，例如 "text_generation" / "chat_completion" / "text_to_image" 等。
        async_mode: 是否按异步客户端解析（一般保持 False，与 infer 用法一致）。
        hf_token: 可选，HF token；不传则从环境变量 HF_TOKEN 读取。

    Returns:
        一个 dict，字段包括：
        - task_type
        - pipeline_tag
        - mapped_method: InferenceClient 上实际调用的方法名
        - required_args: 推理时必须提供的字段名列表
        - output_type_hint: 预期输出类型（text/image/audio/video/embedding/any）
        - client_class: 使用的客户端类名（InferenceClient / AsyncInferenceClient）
        - docstring: 底层方法的文档字符串
        - parameters: 方法签名解析出的参数列表（名称/是否必填/默认值/注解等）
    """
    t0 = time.time()
    try:
        _append_trace(
            {
                "ts": time.time(),
                "tool": "inspect_task",
                "stage": "call",
                "args": _truncate_for_trace(_redact_args(locals())),
            }
        )
        client = ModelInferenceClient(token=hf_token, async_mode=async_mode)
        info = client.get_task_info(task_type=task_type)
        dt_ms = (time.time() - t0) * 1000
        _stat_ok("inspect_task", dt_ms)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "inspect_task",
                "stage": "return",
                "ok": True,
                "dt_ms": dt_ms,
                "result": _truncate_for_trace(info, 1200),
            }
        )
        return info
    except Exception as e:
        dt_ms = (time.time() - t0) * 1000
        _stat_err("inspect_task", dt_ms, e)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "inspect_task",
                "stage": "error",
                "ok": False,
                "dt_ms": dt_ms,
                "error": f"{type(e).__name__}: {str(e)}",
            }
        )
        raise


# =========================
# Tool 5: infer
# =========================

@tool
def infer(
    task_type: str,
    model: str,
    inputs: Any,
    parameters: Optional[Dict[str, JSONScalar]] = None,
    parameters_extra_json: Optional[str] = None,
    provider: Optional[str] = None,
    timeout_s: Optional[float] = None,
    hf_token: Optional[str] = None,
) -> Any:
    """
    调用 Hugging Face Inference Providers 执行推理（真实在线执行）。

    Args:
        task_type: 你的 benchmark task_type。
        model: 要调用的模型 id（Hugging Face model id）。
        inputs: 推理核心输入字段。可以是：
            - 字符串（例如 prompt/text）
            - dict（包含必需字段，如 {"prompt": "..."} / {"text": "..."} / {"image": "..."} 等）
            - bytes（媒体二进制）
        parameters: 可选的推理参数 dict（例如 {"max_new_tokens": 256}），仅支持标量值（str/int/float/bool）。
        parameters_extra_json: 可选。一个 JSON 字符串，解析后必须是 object/dict，用于传递复杂参数（dict/list/嵌套结构）。
            说明：
            - 为了兼容部分 provider 的 function-calling JSON schema 校验，infer 的 `parameters` 仍限制为标量 dict；
              若你需要传递复杂结构（例如 HF 的 `extra_body`、`stop: ["..."]`、`target_size` 等），请使用本字段。
            - 与 `parameters` 同时提供时：先解析 parameters_extra_json，再用 `parameters` 覆盖同名键（标量优先）。
        provider: 可选的 provider 名称（例如 "hf-inference" 或其他 provider，默认auto）。
        timeout_s: 可选的超时秒数（infer 调用级别）。
        hf_token: 可选，HF token；不传则从环境变量 HF_TOKEN 读取。
    Returns:
        - 对文本/JSON/列表：直接返回（可 JSON 序列化）
        - 对 image/audio/video bytes 或 PIL 图像：自动保存到 outputs/assets（或 TOOL_ASSETS_DIR），返回 {"type": "...", "path": "..."}
    """
    t0 = time.time()
    try:
        _append_trace(
            {
                "ts": time.time(),
                "tool": "infer",
                "stage": "call",
                "args": _truncate_for_trace(
                    _redact_args(
                        {
                            "task_type": task_type,
                            "model": model,
                            "provider": provider,
                            "timeout_s": timeout_s,
                            "inputs": inputs,
                            "parameters": parameters,
                            "parameters_extra_json": parameters_extra_json,
                        }
                    )
                ),
            }
        )
        # 工具级硬约束：infer(timeout_s) 上限（对 smolagents / ReAct 一视同仁）
        timeout_max = _hard_limit_infer_timeout_max_s()
        if timeout_max is not None:
            if timeout_s is None:
                timeout_s = float(timeout_max)
            else:
                try:
                    timeout_s = min(float(timeout_s), float(timeout_max))
                except Exception:
                    timeout_s = float(timeout_max)

        # 底层 client：provider/timeout 在这里控制（便于 baseline 统一）
        client = ModelInferenceClient(
            token=hf_token,
            provider=provider,
            timeout=timeout_s,
            async_mode=False,
        )

        # 合并推理参数：
        # - parameters_extra_json 支持复杂结构（dict/list/嵌套）
        # - parameters 仍是标量 dict（用于保持工具 schema 简洁）
        merged_parameters: Dict[str, Any] = {}
        if parameters_extra_json:
            parsed = json.loads(parameters_extra_json)
            if not isinstance(parsed, dict):
                raise ValueError(
                    "parameters_extra_json 必须是 JSON object（例如：{\"stop\": [\"\\n\\n\"]}）。"
                )
            merged_parameters.update(parsed)
        if parameters:
            merged_parameters.update(parameters)

        # 调用推理（不做额外输入规范化，要求调用方显式构造正确的 inputs 字段）
        result = client.inference(
            task_type=task_type,
            inputs=inputs,
            model_id=model,
            parameters=merged_parameters,
        )

        # output_type hint：来自 hf_client 的 TASK_CONFIGS（若没有则 "any"）
        output_type_hint = "any"
        cfg = TASK_CONFIGS.get(task_type)
        if cfg is not None:
            output_type_hint = cfg.output_type

        # 媒体结果落盘 + 分割 mask / 检测·分割可视化叠加（保证可序列化）
        safe_result = prepare_infer_result_for_ui(
            task_type=task_type,
            inputs=inputs,
            raw_result=result,
            output_type_hint=output_type_hint,
            maybe_save_media_fn=_maybe_save_media,
        )

        dt_ms = (time.time() - t0) * 1000
        _stat_ok("infer", dt_ms)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "infer",
                "stage": "return",
                "ok": True,
                "dt_ms": dt_ms,
                # Do not truncate infer() results in tool_trace logs.
                # This is especially important for feature_extraction embeddings, where the
                # string/array representation can be very long and truncation breaks
                # downstream JSON decoding/debugging.
                "result": safe_result,
            }
        )
        return safe_result

    except Exception as e:
        dt_ms = (time.time() - t0) * 1000
        _stat_err("infer", dt_ms, e)
        tb_text = "".join(traceback.format_exception(type(e), e, e.__traceback__))
        _append_trace(
            {
                "ts": time.time(),
                "tool": "infer",
                "stage": "error",
                "ok": False,
                "dt_ms": dt_ms,
                "error": f"{type(e).__name__}: {str(e)}",
                "traceback": tb_text,
            }
        )
        if isinstance(e, StopIteration):
            raise RuntimeError(
                "InferStopIteration: "
                "infer 底层触发 StopIteration；已重包装为普通 RuntimeError，"
                "避免在异步/多 Agent 调度中被 Future 回调链放大。"
            ) from e
        if _looks_like_timeout_error(e):
            raise RuntimeError(
                "InferTimeout: "
                f"infer_timeout_s={timeout_s}; "
                "该错误为实验级超时错误，已捕捉到本次 infer 可能已经超时; "
                f"cause={type(e).__name__}: {str(e)}"
            ) from e
        raise


# =========================
# Tool 6: get_file_info
# =========================

@tool
def get_file_info(
    path: str,
    sheet_name: Optional[str] = None,
    return_table: bool = False,
) -> Dict[str, Any]:
    """
    获取指定路径多模态文件的结构化信息（图片/音频/表格）。

    行为：
    1) 自动根据文件后缀（并辅以 mimetype）判断文件类型。
    2) png/jpg/webp 等图片，wav/mp3 等音频：返回结构化元信息（尽量无依赖可用；可选依赖增强解析）。
    3) csv/tsv/xlsx 等表格：默认仅返回元信息；当 return_table=True 时才返回完整表格内容（二维数组）。

    Args:
        path: 文件路径（相对或绝对均可）。
        sheet_name: 仅对 xlsx 有效；不传则读取第一个 sheet。
        return_table: 仅对表格文件有效。是否返回完整表格内容（二维数组）。默认 False（只返回元信息，避免输出过大）。

    Returns:
        dict，保证可 JSON 序列化：
        {
          "path": "...",
          "exists": true/false,
          "file_kind": "image"|"audio"|"table"|"unknown",
          "suffix": ".png",
          "mime": "image/png",
          "size_bytes": 123,
          "mtime_utc": "2026-03-10T00:00:00+00:00",
          "meta": {...},     # 结构化元信息（按类型不同）
          "table": [...],    # 仅当表格类型且 return_table=True 时返回完整内容（二维数组）
        }
    """
    t0 = time.time()
    try:
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_file_info",
                "stage": "call",
                "args": _truncate_for_trace(
                    {"path": path, "sheet_name": sheet_name, "return_table": return_table}
                ),
            }
        )

        p = os.path.expanduser(path)
        p = os.path.abspath(p)
        exists = os.path.exists(p)
        suffix = os.path.splitext(p)[1].lower()
        mime, _ = mimetypes.guess_type(p)

        base: Dict[str, Any] = {
            "path": p,
            "exists": bool(exists),
            "file_kind": _detect_file_kind_by_suffix(p),
            "suffix": suffix,
            "mime": mime,
            "size_bytes": None,
            "mtime_utc": None,
            "meta": {},
        }

        if not exists:
            dt_ms = (time.time() - t0) * 1000
            _stat_ok("get_file_info", dt_ms)
            _append_trace(
                {
                    "ts": time.time(),
                    "tool": "get_file_info",
                    "stage": "return",
                    "ok": True,
                    "dt_ms": dt_ms,
                    "result": _truncate_for_trace(base, 1200),
                }
            )
            return base

        st = os.stat(p)
        base["size_bytes"] = int(st.st_size)
        base["mtime_utc"] = _ts_iso_utc(float(st.st_mtime))

        kind = base["file_kind"]
        if kind == "image":
            base["meta"] = _image_meta_best_effort(p)
        elif kind == "audio":
            base["meta"] = _audio_meta_best_effort(p)
        elif kind == "table":
            if suffix in (".csv", ".tsv"):
                out = _read_text_table_rows(p, return_table=return_table)
            elif suffix == ".xlsx":
                out = _read_xlsx_table(p, sheet_name=sheet_name, return_table=return_table)
            else:
                # .xls 等：无依赖情况下无法稳定解析，给出结构化提示
                raise ImportError(
                    f"暂不支持解析 {suffix}（建议转为 .csv 或 .xlsx）。"
                )
            base["meta"] = out["meta"]
            if return_table:
                base["table"] = out["table"]
        else:
            # unknown：仅返回基础 stat 信息
            base["meta"] = {"note": "unknown file kind; only basic stat returned"}

        dt_ms = (time.time() - t0) * 1000
        _stat_ok("get_file_info", dt_ms)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_file_info",
                "stage": "return",
                "ok": True,
                "dt_ms": dt_ms,
                "result": _truncate_for_trace(
                    {
                        "file_kind": base.get("file_kind"),
                        "suffix": suffix,
                        "size_bytes": base.get("size_bytes"),
                        "meta": base.get("meta"),
                        "table_rows": len(base.get("table", []) or []),
                    },
                    1200,
                ),
            }
        )
        return base

    except Exception as e:
        dt_ms = (time.time() - t0) * 1000
        _stat_err("get_file_info", dt_ms, e)
        _append_trace(
            {
                "ts": time.time(),
                "tool": "get_file_info",
                "stage": "error",
                "ok": False,
                "dt_ms": dt_ms,
                "error": f"{type(e).__name__}: {str(e)}",
            }
        )
        raise